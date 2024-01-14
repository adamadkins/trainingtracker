import json
import time

import openpyxl
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from datetime import datetime
import random
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import json_normalize
from openpyxl.styles import Font, PatternFill, numbers, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

# Path to the JSON file
JSON_FILE_PATH = 'data.json'
slider_values = 'slider_value.json'


@app.route('/get_slider_value/<slider_id>', methods=['GET'])
def get_slider_value(slider_id):
    data = read_slider_value(slider_id)

    # Check if the entry has 'value' and 'quality_text' keys
    if 'value' in data and 'quality_text' in data:
        return jsonify(value=data['value'], quality_text=data['quality_text'])
    else:
        return jsonify({"error": "Invalid data format"}), 400


@app.route('/save_slider_value/<slider_id>', methods=['POST'])
def save_slider_value(slider_id):
    value = request.json['value']
    write_slider_value(slider_id, value)
    return jsonify(success=True)


@app.route('/run_excel_output')
def run_excel_output():
    # Read data from JSON
    with open('data.json', 'r') as file:
        data = json.load(file)

    # Extract members data
    members_data = data['members']

    # Flatten the JSON data and convert to DataFrame
    df = json_normalize(members_data)

    # Drop the 'id' column if it exists and rename columns
    column_mappings = {
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'start_date': 'Start Date',
        'category': 'Category',
        'overall_progress.register': 'Register',
        'overall_progress.dining_room': 'Dining Room',
        'overall_progress.front_counter_bagging': 'Front Counter Bagging',
        'overall_progress.mobile_bagging': 'Mobile Bagging',
        'overall_progress.ipads': 'iPads',
        'overall_progress.drive_through_bagging': 'Drive-Through Bagging',
        'overall_progress.staging': 'Staging',
        'overall_progress.desserts': 'Desserts',
        'overall_progress.primary_drinks': 'Primary Drinks',
        'overall_progress.secondary_drinks': 'Secondary Drinks',
        'average_percentage': 'Overall Progress'
    }

    df = df.rename(columns=column_mappings).drop(columns=['id'], errors='ignore')

    # Excel file path
    excel_path = 'output.xlsx'

    # Check if the Excel file exists
    if os.path.exists(excel_path):
        # Load the existing workbook
        workbook = openpyxl.load_workbook(excel_path)
        if "Sheet1" in workbook.sheetnames:
            worksheet = workbook["Sheet1"]
            workbook.remove(worksheet)
        worksheet = workbook.create_sheet("Sheet1", 0)
    else:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"

    # Write DataFrame to Excel without the index
    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)

    # Apply header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill

    # Set column widths
    for col in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 3)
        worksheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Format numerical columns as percentages
    percentage_columns = ['Register', 'Dining Room', 'Register',
                          'Dining Room',
                          'Front Counter Bagging',
                          'Mobile Bagging',
                          'iPads',
                          'Drive-Through Bagging',
                          'Staging',
                          'Desserts',
                          'Primary Drinks',
                          'Secondary Drinks',
                          'Overall Progress']
    for col in worksheet.columns:
        if worksheet.cell(row=1, column=col[0].column).value in percentage_columns:
            for cell in col:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00"%"'

    # Freeze header row and apply filters
    worksheet.freeze_panes = 'A2'

    # Save the formatted workbook
    workbook.save(excel_path)

    # Serve the Excel file for download
    return send_file(excel_path, as_attachment=True)


def read_slider_value(slider_id):
    try:
        with open(slider_values, 'r') as file:
            data = json.load(file)
        return data.get(slider_id, {"value": 0, "quality_text": ""})  # Default to 0 and an empty string if not set
    except FileNotFoundError:
        return {"value": 0, "quality_text": ""}


def get_quality_text(slider_value):
    quality_texts = ["Needs Work", "Below Average", "Average", "Above Average", "Excellent"]

    # Convert slider_value to an integer
    slider_value = int(slider_value)

    # Ensure slider_value is within the valid range
    slider_value = max(1, min(len(quality_texts), slider_value))

    return quality_texts[slider_value - 1]


def write_slider_value(slider_id, value):
    try:
        with open(slider_values, 'r') as file:
            data = json.load(file)
    except FileNotFoundError:
        data = {}

    # Get the quality text corresponding to the slider value
    quality_text = get_quality_text(value)

    # Store both the value and quality text
    data[slider_id] = {"value": value, "quality_text": quality_text}

    with open(slider_values, 'w') as file:
        json.dump(data, file, indent=4)  # Indent with 4 spaces for better readability


foh_positions = [
    "ipads", "register", "dining_room", "front_counter_bagging", "mobile_bagging", "drive_through_bagging", "runner",
    "staging",
    'desserts', "primary_drinks", "secondary_drinks"
]

boh_positions = [
    "fries", "buns", "dishes", "nuggets", "boards", "grill",
    'machines', "table"
]


def load_data():
    try:
        with open(JSON_FILE_PATH, 'r') as file:
            data = json.load(file)
    except FileNotFoundError:
        data = {"members": []}

    return data


def save_data(data):
    try:
        # Introduce a delay before writing to the file
        time.sleep(1)  # Delay of 0.5 seconds

        # Write data to the file
        with open('data.json', 'w') as file:
            json.dump(data, file, indent=2)

    except Exception as e:
        print(f"Error writing to data.json: {e}")
        # Handle the error appropriately


@app.route('/update_progress/<int:new_hire_id>/<string:position_id>', methods=['POST'])
def update_progress(new_hire_id, position_id):
    data = load_data()

    new_hire = next((member for member in data["members"] if member["id"] == new_hire_id), None)

    if new_hire:
        new_progress = float(request.form['progress'])

        if position_id == position_id in foh_positions + boh_positions:
            update_overall_progress_in_data(new_hire, position_id, new_progress)
            save_data(data)
            return jsonify({"success": True, "average_progress": new_hire['average_percentage']}), 200
        else:
            return jsonify({"error": f"Invalid position_id: {position_id}", "success": False}), 400
    else:
        return jsonify({"error": "Team member not found", "success": False}), 404


def update_overall_progress_in_data(new_hire, position_id, new_progress):
    new_hire['overall_progress'][position_id] = round(new_progress, 2)

    # Update the average_percentage
    new_hire['average_percentage'] = calculate_average_percentage(new_hire['overall_progress'])


@app.route('/get_overall_progress/<int:new_hire_id>')
def get_overall_progress(new_hire_id):
    data = load_data()

    new_hire = next((member for member in data["members"] if member["id"] == new_hire_id), None)

    if new_hire:
        overall_progress = new_hire['overall_progress']
        return jsonify({"overall_progress": overall_progress}), 200
    else:
        return jsonify({"error": "Team member not found"}), 404


@app.route('/add_new_member', methods=['GET', 'POST'])
def add_new_member():
    data = load_data()

    if request.method == 'POST':
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        start_date = request.form['start_date']
        category = request.form['category']

        # Generate a random 4-digit ID
        new_id = random.randint(1000, 9999)
        used_ids = data.setdefault("used_ids", [])

        # Ensure the generated ID is unique by checking the used_ids list in data.json
        while new_id in used_ids:
            new_id = random.randint(1000, 9999)

        new_member = {
            "id": new_id,
            "first_name": first_name,
            "last_name": last_name,
            "start_date": start_date,
            "category": category,
            "overall_progress": {pos: 0.00 for pos in foh_positions} if category == "FOH" else {pos: 0.00 for pos in
                                                                                                boh_positions},
            "average_percentage": 0.00,
        }

        data["members"].append(new_member)

        # Add the new ID to the used_ids list in data.json
        data["used_ids"].append(new_id)

        save_data(data)

        return redirect(url_for('index'))

    return render_template('index.html')


@app.route('/delete_member/<int:member_id>', methods=['DELETE'])
def delete_member(member_id):
    # Load data at the beginning of the route
    data = load_data()

    # Retrieve the team member from your data
    member = next((m for m in data['members'] if m['id'] == member_id), None)

    if member:
        # Remove the team member from your data
        data['members'].remove(member)

        # Save the updated data
        save_data(data)

        return jsonify({"success": True}), 200
    else:
        return jsonify({"error": "Team member not found", "success": False}), 404


@app.route('/templates/index.html')
def index():
    data = load_data()
    for member in data["members"]:
        # Convert start_date string to datetime object
        start_date = datetime.strptime(member["start_date"], "%Y-%m-%d")

        # Add formatted date to the member dictionary
        member["formatted_start_date"] = start_date.strftime('%B %d, %Y')
    foh_members = [member for member in data["members"] if member["category"] == 'FOH']
    boh_members = [member for member in data["members"] if member["category"] == 'BOH']

    foh_training_positions = [pos for pos in foh_positions]
    boh_training_positions = [pos for pos in boh_positions]

    overall_progress_dict = {}
    for member in foh_members + boh_members:
        rounded_percentage = round(calculate_average_percentage(member["overall_progress"]), 2)
        overall_progress_dict[member["id"]] = {
            'average_percentage': rounded_percentage,
            'positions': member["overall_progress"]
        }

    return render_template('index.html', foh_members=foh_members, boh_members=boh_members,
                           foh_training_positions=foh_training_positions, boh_training_positions=boh_training_positions,
                           overall_progress_dict=overall_progress_dict, enumerate=enumerate)


def calculate_average_percentage(overall_progress):
    values = [value for value in overall_progress.values() if value is not None]
    if values:
        return sum(values) / len(values)
    else:
        return 0.0


if __name__ == "__main__":
    app.debug = True
    app.run(host='0.0.0.0')
